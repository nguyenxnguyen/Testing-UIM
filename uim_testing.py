import openpyxl as opx
import itertools
import os
import mmap
import tkFileDialog
import tkMessageBox
from Tkinter import *
import ttk
import ScrolledText
import time
import xml.etree.ElementTree as ET


def timeit(method):
    def timed(*args, **kw):
        ts = time.time()
        result = method(*args, **kw)
        te = time.time()
        print '%r %2.2f sec' % (method.__name__, te-ts)
        return result
    return timed


class Testing(Frame):
    def __init__(self, master, *args, **kwargs):
        Frame.__init__(self, master, *args, **kwargs)

    @timeit
    def read_excel_detail(self, file_path):
        wb = opx.load_workbook(file_path, data_only=True)
        sh_oid = wb.get_sheet_by_name('OIDs_List')
        rows_oid = sorted(list(itertools.islice(sh_oid.iter_rows(), 1, None)))
        info_oid = {}
        for row in rows_oid:
            if not row:
                continue
            oid_name = str(row[0].value)
            oid_value = str(row[1].value)
            if check_document.get() == 'input':
                iskey = str(row[2].value)
                needsdelta = str(row[3].value)
                oid_type = str(row[4].value)
                info_oid[oid_name] = [oid_value, iskey, needsdelta, oid_type]
            else:
                info_oid[oid_name] = [oid_value]

        sh_detail = wb.get_sheet_by_name('Detail')
        if check_document.get() == 'input':
            sh_column = ['Vendor Cert', 'Vendor Cert Name', 'Metric Family', 'Metric Family Name',
                         'New MF Introduced  -Yes/No', 'New VC Introduced  -Yes/No', 'VendorPriority', 'Metric',
                         'New Metrics \n - Yes/No', 'Expression', 'QOS Name', 'New OID  - Yes/No',
                         'Data Type', 'Rollup Strategy', 'Unit', 'Sim', 'Remarks']
        else:
            sh_column = ['Vendor Cert', 'Vendor Cert Name', 'Metric Family', 'Metric Family Name',
                         'New MF Introduced -Yes/No', 'New VC Introduced -Yes/No', 'VendorPriority', 'Metric',
                         'New Metrics - Yes / No', 'Expression', 'QOS Name', 'New OID  - Yes/No',
                         'Data Type', 'Rollup Strategy', 'Unit', 'Sim', 'Remarks']
        msg_warning = ''
        for i in xrange(0, 17):
            if sh_detail['1'][i].value != sh_column[i]:
                msg_warning += '\n----------\n'
                msg_warning += 'Column: %s\nMust be: %s' % (sh_detail['1'][i].value, sh_column[i])

        if msg_warning:
            #tkMessageBox.showerror("Wrong Format", msg_warning)
            scroll_text.insert(INSERT, msg_warning)

        col_vc = 'A'
        col_mf = 'C'
        col_new_mf = 'E'
        col_new_vc = 'F'
        col_metric = 'H'
        col_new_metric = 'I'
        col_exp = 'J'
        col_data_type = 'M'
        col_rollup = 'N'
        col_unit = 'O'
        rows = sorted(list(itertools.islice(sh_detail.iter_rows(), 1, None)))
        info_vc = {}
        info_mf = {}
        list_vc = []
        for row in rows:
            vc_name = str(row[0].value)
            if vc_name and vc_name not in list_vc:
                list_vc.append(vc_name)
        for vc_name in list_vc:
            info_metric = {}
            for row in rows:
                if row[0].value != vc_name:
                    continue
                mf_name = str(row[2].value)
                metric = str(row[7].value)
                exp = str(row[9].value)
                if row[12].value is None:
                    data_type = 'data_type_not_defined'
                else:
                    data_type = str(row[12].value)
                if row[13].value is None:
                    rollup = 'rollup_not_defined'
                else:
                    rollup = str(row[13].value)
                if row[14].value is None:
                    unit = 'unit_not_defined'
                else:
                    unit = str(row[14].value)
                info_metric[metric] = [exp, data_type, rollup, unit]
                info_mf[mf_name] = ['mf_not_found', 'mm_not_defined', 'vp_not_defined']
                info_vc[vc_name] = ['vc_not_found', mf_name, info_metric]
        return info_vc, info_mf, info_oid, msg_warning

    @staticmethod
    def get_input():
        input_entry.delete(0, END)
        file_path = tkFileDialog.askopenfilename(initialdir=os.path.expanduser('~/Desktop'),
                                                 title='Input or Document file')
        if 'Input' in os.path.split(file_path)[1]:
            check_document.set('input')
        else:
            check_document.set('document')
        input_entry.insert(INSERT, file_path)

    @staticmethod
    def get_folder():
        folder_entry.delete(0, END)
        folder = tkFileDialog.askdirectory(initialdir=os.path.expanduser('~/Desktop'),
                                           title='Where the root dir of XMLs?')
        folder_entry.insert(INSERT, folder)

    @staticmethod
    @timeit
    def find_xml(info_vc, info_mf):
        folder = folder_entry.get()
        #folder = 'y:/network/snmpcollector/'
        if not folder:
            folder = 'y:/network/snmpcollector/'
            #folder = 'd:/Dropbox/Temp/'
            #folder = 'v:/'

        vc_folder = os.path.join(folder, 'VendorCertifications/')
        mf_folder = os.path.join(folder, 'MetricFamily/')
        mm_folder = os.path.join(folder, 'Publish/')
        vp_folder = os.path.join(folder, 'VendorPriorities/')
        f_stream = {}
        ts = time.time()
        i = 0
        c = 0
        missing_msg = ''
        vc_found = 0
        for vc_file in os.listdir(vc_folder):
            if vc_file.endswith('.xml'):
                full_path = os.path.join(vc_folder, vc_file)
                f_open = open(full_path, 'r')
                s = mmap.mmap(f_open.fileno(), 0, access=mmap.ACCESS_READ)
                found_flag = 0
                for vc_key in info_vc.keys():
                    i += 1
                    vc_key_facet = '=\"%s\"' % vc_key
                    if s.find(vc_key_facet) != -1:
                        if info_vc[vc_key][0] != 'vc_not_found':
                            missing_msg += 'Duplicate VC %s: %s | %s\n' % (vc_key, info_vc[vc_key][0], vc_file)
                        elif s.find(info_vc[vc_key][1]) != -1:
                            vc_found += 1
                            found_flag = 1
                            info_vc[vc_key][0] = vc_file
                            f_stream[vc_file] = f_open
                        else:
                            vc_found += 1
                            missing_msg += 'Wrong info VC: %s - MF: %s\n' % (vc_key, info_vc[vc_key][1])
                s.close()
                if found_flag == 0:
                    c += 1
                    f_open.close()
                if vc_found == len(info_vc):
                    break
        if True:
            for vc_key in info_vc.keys():
                if info_vc[vc_key][0] == 'vc_not_found':
                    missing_msg += 'Missing VC: %s - MF: %s\n' % (vc_key, info_vc[vc_key][1])
                    info_vc.pop(vc_key)

        print 'Done find VC %s' % (time.time() - ts)
        print f_stream
        mf_found = 0
        for mf_file in os.listdir(mf_folder):
            if mf_file.endswith('.xml'):
                full_path = os.path.join(mf_folder, mf_file)
                f_open = open(full_path, 'r')
                s = mmap.mmap(f_open.fileno(), 0, access=mmap.ACCESS_READ)
                found_flag = 0
                for mf_key in info_mf.keys():
                    if info_mf[mf_key][0] != 'mf_not_found':
                        continue
                    if s.find(mf_key) != -1:
                        info_mf[mf_key][0] = mf_file
                        found_flag = 1
                        mf_found += 1
                        f_stream[mf_file] = f_open
                s.close()
                if found_flag == 0:
                    c += 1
                    f_open.close()
                if mf_found == len(info_mf):
                    break

        print 'Done find MF %s' % (time.time() - ts)

        # --- find Metric Map ---
        mm_found = 0
        for mm_file in os.listdir(mm_folder):
            if mm_file.endswith('.xml'):
                full_path = os.path.join(mm_folder, mm_file)
                for mf_key in info_mf.keys():
                    if info_mf[mf_key][1] != 'mm_not_defined':
                        continue
                    if mf_key in mm_file:
                        mm_found += 1
                        info_mf[mf_key][1] = mm_file
                        f_open = open(full_path, 'r')
                        f_stream[mm_file] = f_open
                        break
                if mm_found == len(info_mf):
                    break

        print 'Done find MM %s' % (time.time() - ts)

        # --- find Vendor Priorities ---
        vp_found = 0
        for vp_file in os.listdir(vp_folder):
            if vp_file.endswith('.xml'):
                full_path = os.path.join(vp_folder, vp_file)
                f_open = open(full_path, 'r')
                s = mmap.mmap(f_open.fileno(), 0, access=mmap.ACCESS_READ)
                found_flag = 0
                for mf_key in info_mf.keys():
                    if info_mf[mf_key][2] != 'vp_not_defined':
                        continue
                    if s.find(mf_key) != -1:
                        info_mf[mf_key][2] = vp_file
                        found_flag = 1
                        vp_found += 1
                        f_stream[vp_file] = f_open
                        for vc_key in info_vc.keys():
                            if info_vc[vc_key][1] == mf_key:
                                if s.find(vc_key) == -1:
                                    missing_msg += 'Missing VC: %s in VP: %s\n' % (vc_key, vp_file)
                        break
                s.close()
                if found_flag == 0:
                    c += 1
                    f_open.close()
                if vp_found == len(info_mf):
                    break

        print 'Done find VP %s' % (time.time() - ts)

        # --- check if there are some missing file ---
        for mf_key in info_mf:
            if info_mf[mf_key][0] == 'mf_not_found':
                missing_msg += '%s is not found\n' % mf_key
            if info_mf[mf_key][1] == 'mm_not_defined':
                missing_msg += 'MM for %s is not found\n' % mf_key
            if info_mf[mf_key][2] == 'vp_not_defined':
                missing_msg += 'VP for %s is not found\n' % mf_key

        print 'Loop: %s' % i
        #print 'Closed files: %s' % c
        if missing_msg:
            scroll_text.insert(INSERT, missing_msg)

        return info_vc, info_mf, f_stream, missing_msg

    @staticmethod
    def read_xml_mf(info_mf, mf_key, f_stream):
        list_non_polled = ['Indexes', 'Names', 'Descriptions', 'SourceFacetTypes', 'ItemID']
        list_allowed_type = ['double', 'int', 'long']
        xml_info_metric = {}
        msg = ''
    # --- Read MetricFamily files ---
        mf_file = info_mf[mf_key][0]
        mf_open = f_stream[mf_file]
        tree = ET.ElementTree(ET.fromstring(mf_open.read()))
        root_mf = tree.getroot()
        mf_facets = root_mf.findall('FacetType')
        mf_info = {}
        for mf_facet in mf_facets:
            if mf_facet.attrib.get('name') == mf_key:
                list_of_attribute = []
                list_of_attribute.extend(mf_facet.findall('Attribute'))
                mf_attribute_groups = mf_facet.findall('AttributeGroup')
                for mf_attribute_group in mf_attribute_groups:
                    list_of_attribute.extend(mf_attribute_group.findall('Attribute'))
                for mf_attribute in list_of_attribute:
                    metric_name = mf_attribute.attrib.get('name')
                    metric_type = mf_attribute.attrib.get('type')
                    if metric_type.lower() not in list_allowed_type:
                        #print 'SKIP: %s - %s' % (metric_name, metric_type.lower())
                        metric_rollup = 'rollup_not_defined'
                    elif metric_name in list_non_polled or mf_attribute.find('RollupStrategy') is None:
                        metric_rollup = 'rollup_not_defined'
                    else:
                        metric_rollup = mf_attribute.find('RollupStrategy').text
                    mf_info[metric_name] = [metric_type, metric_rollup]
            else:
                continue

    # --- Read MetricMap files ---
        mm_file = info_mf[mf_key][1]
        mm_open = f_stream[mm_file]
        tree = ET.ElementTree(ET.fromstring(mm_open.read()))
        root_mm = tree.getroot()
        mm_def = root_mm.find('MetricFamilyDef')
        ci_type = mm_def.attrib.get('ciType')
        ci_descr = mm_def.attrib.get('name')
        polled_list = mm_def.find('PolledMetrics')
        polled_metrics = polled_list.findall('PolledMetricDef')

        mm_info = {}
        for metric in polled_metrics:
            metric_name = metric.attrib.get('name')
            metric_type = metric.attrib.get('type')
            metric_unit = metric.attrib.get('units')
            metric_id = metric.attrib.get('ciMetricId')
            mm_info[metric_name] = [metric_id, metric_type, metric_unit]

        non_polled_list = mm_def.find('NonPolledMetrics')
        non_polled_metrics = non_polled_list.findall('NonPolledMetricDef')
        for metric in non_polled_metrics:
            metric_name = metric.attrib.get('name')
            metric_type = metric.attrib.get('type')
            metric_unit = 'unit_not_defined'
            metric_id = 'metric_id_not_defined'
            mm_info[metric_name] = [metric_id, metric_type, metric_unit]

        list_key = []
        for key in mf_info.keys():
            if mf_info[key][0].lower() not in list_allowed_type:
                if key in mm_info.keys():
                    pass
                else:
                    continue
            if key not in mm_info.keys():
                msg += 'Metric %s of %s not defined in %s\n' % (key, mf_key, mm_file)
                #scroll_text.insert(INSERT, msg)
            else:
                list_key.append(key)
        if list_key:
            for key in list_key:
                data_type = mm_info[key][1]
                rollup = mf_info[key][1]
                unit = mm_info[key][2]
                xml_info_metric[key] = [data_type, rollup, unit]

    # --- Read VendorPriorities files ---
        vp_file = info_mf[mf_key][2]
        vp_open = f_stream[vp_file]
        tree = ET.ElementTree(ET.fromstring(vp_open.read()))
        root_vp = tree.getroot()
        mf_id = root_vp.find('MetricFamilyID').text
        id_name = mf_id.replace('{http://im.ca.com/normalizer}', '')
        vp_info = []
        if id_name == mf_key:
            cert_order_list = root_vp.find('CertificationOrderList')
            cert_orders = cert_order_list.findall('CertificationOrder')
            for cert_order in cert_orders:
                vc = cert_order.find('VendorCertID').text
                vc = vc.replace('{http://im.ca.com/certifications/snmp}', '')
                vp_info.append(vc)
        duplicate_vc = set([vc for vc in vp_info if vp_info.count(vc) > 1])
        if duplicate_vc:
            msg_dup = ''
            for duplicate in duplicate_vc:
                msg_dup += ' - %s' % duplicate
            msg += 'Duplicates in %s: %s\n' % (vp_file, msg_dup)
        return xml_info_metric, vp_info, msg

    @staticmethod
    def read_xml_vc(vc_key, vc_file, f_stream, xml_info_mf, xml_info_vp):
        xml_oid_vc_extra = {}
        xml_info_vc_extra = {}
        msg = ''

        # --- Read VendorCert files ---
        vc_open = f_stream[vc_file]
        tree = ET.ElementTree(ET.fromstring(vc_open.read()))
        root_vc = tree.getroot()
        facet_types = root_vc.findall('FacetType')
        for facet in facet_types:
            if facet.attrib.get('name') == vc_key:
                attribute_groups = facet.findall('AttributeGroup')
                xml_oid_extra = {}
                for group in attribute_groups:
                    attributes = group.findall('Attribute')
                    for att in attributes:
                        oid = att.find('Source').text
                        oid_name = att.attrib.get('name')
                        oid_type = att.attrib.get('type')
                        is_key = att.find('IsKey').text
                        if att.find('IsIndex') is not None:
                            is_index = att.find('IsIndex').text
                        else:
                            is_index = 'not_defined'
                        if att.find('NeedsDelta') is not None:
                            needs_delta = att.find('NeedsDelta').text
                        else:
                            needs_delta = 'not_defined'
                        xml_oid_extra[oid_name] = [oid, is_key, needs_delta, oid_type, is_index]
                xml_oid_vc_extra[vc_key] = xml_oid_extra
                expressions = facet.find('Expressions')
                expression_group = expressions.find('ExpressionGroup')
                mf_name = expression_group.attrib.get('destCert')
                mf_name = mf_name.replace('{http://im.ca.com/normalizer}', '')
                vp_found = 0
                for vp_info in xml_info_vp[mf_name]:
                    if vp_info == vc_key:
                        vp_found = 1
                        break
                if vp_found == 0:
                    msg += 'Missing prioritiy VC: %s in MF: %s\n' % (vc_key, mf_name)
                    #scroll_text.insert(INSERT, msg)
                list_expression = expression_group.findall('Expression')
                list_metrics = {}
                for expression in list_expression:
                    metric = expression.attrib.get('destAttr')
                    if metric == 'Indexes':
                        continue

                    if metric == 'Names' or metric == 'Descriptions':
                        data_type = xml_info_mf[mf_name][metric][0]
                        rollup = 'rollup_not_defined'
                        unit = 'unit_not_defined'
                    elif metric in xml_info_mf[mf_name].keys():
                        data_type = xml_info_mf[mf_name][metric][0]
                        rollup = xml_info_mf[mf_name][metric][1]
                        unit = xml_info_mf[mf_name][metric][2]
                    else:
                        msg += 'Metric %s of %s is missing from MF %s\n' % (metric, vc_key, mf_name)
                        data_type = 'data_type_is_missing'
                        rollup = 'rollup_is_missing'
                        unit = 'unit_is_missing'
                        #scroll_text.insert(INSERT, msg)

                    metric_exp = expression.text
                    metric_exp = metric_exp.replace('&&', '&amp;&amp;')
                    list_metrics[metric] = [metric_exp, data_type, rollup, unit]

                xml_info_vc_extra = [vc_file, mf_name, list_metrics]
            else:
                continue

        vc_open.close()
        return xml_oid_vc_extra, xml_info_vc_extra, msg

    def read_xml(self, info_vc, info_mf, f_stream):
        xml_oid_vc = {}
        xml_info_mf = {}
        xml_info_vc = {}
        xml_info_vp = {}
        prefix_msg = '\n\nWARNING FOR XMLs\n\n'
        msg_full = ''
        for mf_key in info_mf.keys():
            xml_info_metric, vp_info, msg = self.read_xml_mf(info_mf, mf_key, f_stream)
            xml_info_mf[mf_key] = xml_info_metric
            xml_info_vp[mf_key] = vp_info
            if msg:
                msg_full += msg
        for vc_key in info_vc.keys():
            vc_file = info_vc[vc_key][0]
            xml_oid_vc_extra, xml_info_metric, msg = self.read_xml_vc(vc_key, vc_file, f_stream,
                                                                      xml_info_mf, xml_info_vp)
            xml_oid_vc.update(xml_oid_vc_extra)
            xml_info_vc[vc_key] = xml_info_metric
            if msg:
                msg_full += msg
        #print xml_oid
        #print info_vc['NetscalerGSLBSitesMib']
        #print xml_info_vc['NetscalerGSLBSitesMib']
        if msg_full:
            msg_full = '%s%s' % (prefix_msg, msg_full)
            scroll_text.insert(INSERT, msg_full)
        return xml_info_vc, xml_oid_vc

    def process(self):
        msg_metric_temp_full = ''
        scroll_text.delete(0.0, END)
        while not input_entry.get():
            self.get_input()
        while not folder_entry.get():
            self.get_folder()
        file_path = input_entry.get()
        info_vc, info_mf, info_oid, msg_warning = self.read_excel_detail(file_path)
        if msg_warning:
            pass
        else:
            info_vc, info_mf, f_stream, missing_msg = self.find_xml(info_vc, info_mf)
            #print info_vc
            #print info_mf
            #print f_stream.keys()
            if missing_msg:
                pass
            else:
                pass
                #print f_stream.keys()
            xml_info_vc, xml_oid_vc = self.read_xml(info_vc, info_mf, f_stream)
            #print xml_info_vc
            msg_vc = ''
            for vc in xml_info_vc.keys():
                msg_vc += '\n\t' + vc
            for vc_key in info_vc.keys():
                msg_mf = ''
                msg_oid = ''
                if info_vc[vc_key][1] != xml_info_vc[vc_key][1]:
                    msg_mf = 'MF of %s: %s | %s\n' % (vc_key, info_vc[vc_key], xml_info_vc[vc_key])
                msg_metric_missing = ''
                msg_metric = ''
                for metric_key in info_vc[vc_key][2].keys():
                    if metric_key not in xml_info_vc[vc_key][2].keys():
                        msg_metric_missing += '\nMetric %s of VC %s is missing from XML\n' % (metric_key, vc_key)
                    else:
                        error = 0
                        msg_metric_temp = ''

                        exp = info_vc[vc_key][2][metric_key][0]
                        exp_xml = xml_info_vc[vc_key][2][metric_key][0]
                        if exp != exp_xml:
                            msg_metric_temp += '\t  %s\n\t  | %s\n' % (exp, exp_xml)
                            error += 1
                        data_type = info_vc[vc_key][2][metric_key][1]
                        data_type_xml = xml_info_vc[vc_key][2][metric_key][1]
                        if data_type != data_type_xml:
                            if metric_key == 'Names' or metric_key == 'Descriptions':
                                if data_type == 'String' or data_type is None or data_type == 'data_type_not_defined':
                                    pass
                                else:
                                    msg_metric_temp += '\t  %s\n\t  | %s\n' % (data_type, data_type_xml)
                                    error += 1
                            else:
                                msg_metric_temp += '\t  %s\n\t  | %s\n' % (data_type, data_type_xml)
                                error += 1
                        rollup = info_vc[vc_key][2][metric_key][2]
                        rollup_xml = xml_info_vc[vc_key][2][metric_key][2]
                        if rollup != rollup_xml:
                            msg_metric_temp += '\t  %s\n\t  | %s\n' % (rollup, rollup_xml)
                            error += 1
                        unit = info_vc[vc_key][2][metric_key][3]
                        unit_xml = xml_info_vc[vc_key][2][metric_key][3]
                        if unit != unit_xml:
                            msg_metric_temp += '\t  %s\n\t  | %s\n' % (unit, unit_xml)
                            error += 1
                        if error:
                            msg_metric += '\n\t%s\n%s' % (metric_key, msg_metric_temp)
                        else:
                            pass
                index_found = 0
                for xml_oid in xml_oid_vc[vc_key].keys():
                    if xml_oid == 'Index':
                        index_oid = xml_oid_vc[vc_key][xml_oid][0]
                        index_found = 1
                    msg_oid_temp = ''
                    if check_document.get() == 'input':
                        for oid in info_oid:
                            if oid == xml_oid:
                                oid_value = info_oid[oid][0]
                                oid_iskey = info_oid[oid][1]
                                oid_needsdelta = info_oid[oid][2]
                                oid_type = info_oid[oid][3]

                                if ' ' in oid_value:
                                    msg_oid_temp += '\t  OID in Spreadsheet has space\n'

                                xml_oid_value = xml_oid_vc[vc_key][oid][0]
                                xml_oid_iskey = xml_oid_vc[vc_key][oid][1]
                                xml_oid_needsdelta = xml_oid_vc[vc_key][oid][2]
                                xml_oid_type = xml_oid_vc[vc_key][oid][3]

                                if ' ' in xml_oid_value:
                                    msg_oid_temp += '\t  OID in VC File has space\n'

                                if oid_value != xml_oid_value:
                                    msg_oid_temp += '\t  Value: \t%s\n\t\t  | %s' % (oid_value, xml_oid_value)
                                if oid_iskey != xml_oid_iskey:
                                    msg_oid_temp += '\t  IsKey: \t%s\n\t\t  | %s' % (oid_iskey, xml_oid_iskey)
                                if oid_needsdelta != xml_oid_needsdelta:
                                    msg_oid_temp += '\t  Delta: \t%s\n\t\t  | %s' % (oid_needsdelta, xml_oid_needsdelta)
                                if oid_type != xml_oid_type:
                                    msg_oid_temp += '\t  Type:  \t%s\n\t\t  | %s' % (oid_type, xml_oid_type)
                    else:
                        for oid in info_oid:
                            if oid == xml_oid:
                                oid_value = info_oid[oid][0]
                                if ' ' in oid_value:
                                    msg_oid_temp += '\t  OID in Spreadsheet has space\n'
                                xml_oid_value = xml_oid_vc[vc_key][oid][0]
                                if ' ' in xml_oid_value:
                                    msg_oid_temp += '\t  OID in VC File has space\n'
                                if oid_value != xml_oid_value:
                                    msg_oid_temp += '\t  Value: \t%s\n\t\t  | %s' % (oid_value, xml_oid_value)

                    if msg_oid_temp:
                        msg_oid += '\n\tOID: %s\n%s\n' % (xml_oid, msg_oid_temp)
                if index_found:
                    for xml_oid in xml_oid_vc[vc_key].keys():
                        if xml_oid == 'Index':
                            continue
                        if index_oid == xml_oid_vc[vc_key][xml_oid][0]:
                            index_found = 10
                            break
                if index_found == 1:
                    msg_oid += '\n\tOID Index is different from Other OIDs - %s\n' % index_oid

                if msg_metric or msg_oid:
                    msg_metric_temp_full += '\n%s' % vc_key
                    if msg_metric:
                        msg_metric_temp_full += '%s%s%s' % (msg_mf, msg_metric_missing, msg_metric)
                    if msg_oid:
                        msg_metric_temp_full += msg_oid

            if msg_metric_temp_full:
                msg_metric_full = 'VC Found: ' + msg_vc + \
                                  '\n------------------------------------\n' + \
                                  'MISMATCH INFORMATION between SPREADSHEET AND XMLs\n' + msg_metric_temp_full
            else:
                msg_metric_full = 'VC Found: ' + msg_vc + \
                                  '\n------------------------------------\n' + \
                                  'NO ERRORS DETECTED\n'

            scroll_text.insert(INSERT, msg_metric_full)


if __name__ == "__main__":
    root = Tk()
    root.resizable(width=False, height=False)
    root.title("Check Code")
    note_book = ttk.Notebook(root)
    page1 = ttk.Frame(note_book)
    #page2 = ttk.Frame(note_book)
    note_book.add(page1, text='General')
    #note_book.add(page2, text='Multi-Port')
    note_book.pack(expand=1, fill="both")
    #note_book.select(0)

    test = Testing(root)

    #----------------- PAGE 1 --------------------
    frame11 = Frame(page1)
    frame11.grid(column=1, row=1)
    frame12 = Frame(page1)
    frame12.grid(column=1, row=2)

    blank_label = Label(frame11, text="", width=5)
    blank_label.grid(column=1, row=1)

    check_document = StringVar()
    check_document.set('input')
    document_radio = Radiobutton(frame11, text="Document", variable=check_document, value='document', width=8)
    document_radio.grid(column=2, row=2, sticky=W)

    input_radio = Radiobutton(frame11, text="Input", variable=check_document, value='input', width=8)
    input_radio.grid(column=3, row=2, sticky=W)

    input_entry = Entry(frame11, width=40)
    input_entry.grid(column=4, row=2)
    blank_label = Label(frame11, text="", width=1)
    blank_label.grid(column=5, row=2)

    # Browse input button
    browse_input_button = Button(frame11, text="Browse", command=test.get_input)
    browse_input_button.grid(column=5, row=2, sticky=E)

    blank_label = Label(frame11, text="", width=5)
    blank_label.grid(column=6, row=2)

    folder_label = Label(frame11, text="XMLs Folder:", width=10)
    folder_label.grid(column=7, row=2)
    folder_entry = Entry(frame11, width=40)
    folder_entry.grid(column=8, row=2)
    blank_label = Label(frame11, text="", width=1)
    blank_label.grid(column=9, row=2)

    # Browse xml folder button
    browse_xml_button = Button(frame11, text="Browse", command=test.get_folder)
    browse_xml_button.grid(column=10, row=2, sticky=E)

    blank_label = Label(frame11, text="", width=5)
    blank_label.grid(column=11, row=2)

    blank_label = Label(frame12, text="", width=5)
    blank_label.grid(column=1, row=1)
    blank_label = Label(frame12, text="", width=5)
    blank_label.grid(column=3, row=1)

    # Run button
    run_button = Button(frame12, text="Process", command=test.process, width=10)
    run_button.grid(column=2, row=2)

    blank_label = Label(frame12, text="SPREADSHEET")
    blank_label.grid(column=2, row=3, sticky=W)

    blank_label = Label(frame12, text="|  XML")
    blank_label.grid(column=2, row=4, sticky=W)

    scroll_text = ScrolledText.ScrolledText(frame12, width=120, height=40)
    scroll_text.grid(column=2, row=5)

    root.mainloop()